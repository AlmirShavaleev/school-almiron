from __future__ import annotations

import hashlib
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


COURSE_ROOT = Path(r"D:\На сайт\ЕГЭ Математика")
OUT_DIR = Path(r"D:\Школа Almiron\outputs\course_audit")
XLSX_PATH = OUT_DIR / "Аудит_материалов_курса.xlsx"
MD_PATH = OUT_DIR / "Аудит_материалов_курса_отчет.md"

TEXT_EXTS = {".pdf", ".doc", ".docx", ".ppt", ".pptx", ".txt", ".rtf", ".xlsx", ".xls"}
VIDEO_EXTS = {".mp4", ".mov", ".avi", ".mkv", ".webm", ".m4v"}
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".svg"}
ARCHIVE_EXTS = {".zip", ".rar", ".7z", ".tar", ".gz"}


def norm_text(value: str) -> str:
    value = value.lower().replace("_", " ")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def normalize_for_duplicate(value: str) -> str:
    value = norm_text(Path(value).stem)
    value = re.sub(r"\(\d+\)$", "", value).strip()
    value = re.sub(r"\b[0-9a-zа-я]{5,7}$", "", value, flags=re.I).strip()
    value = re.sub(r"[^0-9a-zа-яё№]+", " ", value, flags=re.I).strip()
    return re.sub(r"\s+", " ", value)


def win_link(path: Path) -> str:
    absolute = str(path.absolute())
    return "file:///" + quote(absolute.replace("\\", "/"), safe="/:()#№+,-._ ")


def long_path(path: Path) -> str:
    absolute = str(path.absolute())
    if os.name == "nt" and not absolute.startswith("\\\\?\\"):
        return "\\\\?\\" + absolute
    return absolute


def rel_parts(path: Path) -> list[str]:
    return list(path.relative_to(COURSE_ROOT).parts)


def extract_topic_number(name: str) -> str:
    patterns = [
        r"^\s*(\d{1,3}(?:[.,-]\d{1,3})?)\s*[.)]",
        r"№\s*([0-9]{1,2}(?:\s*[-,]\s*[0-9]{1,2})*)",
        r"\bвариант\s*№?\s*([0-9]{1,3})",
        r"\bвебинар\s*([0-9]{1,3})",
        r"\bдз\s*-?\s*([0-9]{1,3})",
    ]
    for pattern in patterns:
        match = re.search(pattern, name, flags=re.I)
        if match:
            return re.sub(r"\s+", "", match.group(1))
    return ""


def clean_topic_title(name: str) -> str:
    title = re.sub(r"^\s*\d{1,3}\s*[.)]\s*", "", name).strip()
    title = re.sub(r"^\s*\d{1,2}\s+\d{1,3}[.)]\s*", "", title).strip()
    return title


def material_type(file_path: Path) -> str:
    text = norm_text(" ".join(file_path.parts[-4:]))
    suffix = file_path.suffix.lower()

    if any(word in text for word in ["решения", "решение", "разбор", "разборы"]):
        return "решение домашнего задания"
    if re.search(r"(^|[\s_])дз($|[\s_\-])", text) or any(
        word in text for word in ["домашняя", "домашнее", "домашка", "рабочая тетрадь", "тетрадь"]
    ):
        return "домашнее задание"
    if any(word in text for word in ["ответ", "ключи"]):
        return "ответы"
    if any(word in text for word in ["презентац", "слайды"]):
        return "презентация"
    if any(word in text for word in ["конспект", "шпаргалка", "памятка"]):
        return "конспект"
    if any(word in text for word in ["теория", "теоретич"]):
        return "теория"
    if any(word in text for word in ["вебинар", "урок", "эфир", "запись"]):
        return "вебинар"
    if any(word in text for word in ["самостоятельная", "сам работа", "ср"]):
        return "самостоятельная работа"
    if any(word in text for word in ["контрольная", "кр"]):
        return "контрольная работа"
    if any(word in text for word in ["доп", "дополнитель", "прототип", "вариант"]):
        return "дополнительные задачи"
    if suffix in VIDEO_EXTS:
        return "запись урока"
    if suffix in IMAGE_EXTS:
        return "изображения и вспомогательные материалы"
    if suffix in ARCHIVE_EXTS:
        return "архив"
    if suffix in TEXT_EXTS:
        return "теория"
    return "другие учебные файлы"


def topic_key_for_file(file_path: Path) -> Path:
    return file_path.parent


def bool_yes(value: bool) -> str:
    return "Да" if value else "Нет"


def file_hash(path: Path) -> str:
    h = hashlib.sha1()
    with open(long_path(path), "rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def collect_tree():
    files = []
    dirs = []
    inaccessible = []
    for current, dirnames, filenames in os.walk(COURSE_ROOT):
        current_path = Path(current)
        dirs.append(current_path)
        for filename in filenames:
            path = current_path / filename
            try:
                stat = os.stat(long_path(path))
            except OSError as exc:
                inaccessible.append((path, str(exc)))
                continue
            files.append({"path": path, "stat": stat})
    return files, dirs, inaccessible


def topic_meta(topic_path: Path) -> dict:
    parts = rel_parts(topic_path)
    section = parts[0] if len(parts) >= 1 else COURSE_ROOT.name
    module = parts[1] if len(parts) >= 2 else ""
    topic_name = parts[-1] if parts else COURSE_ROOT.name
    number = extract_topic_number(topic_name) or extract_topic_number(module)
    return {
        "section": section,
        "module": module,
        "number": number,
        "title": clean_topic_title(topic_name),
        "path": topic_path,
    }


def classify_topic_status(flags: dict, count_files: int, missing: list[str], uncertain: bool) -> str:
    has_main = flags["theory"] or flags["conspect"] or flags["presentation"] or flags["video"]
    has_hw = flags["homework"]
    has_solution = flags["solution"]
    if count_files == 0:
        return "Пустая тема"
    if uncertain:
        return "Требует ручной проверки"
    if has_main and has_hw and has_solution and not missing:
        return "Полный комплект"
    if has_main and not has_hw and not has_solution:
        return "Только теория"
    if has_hw and not has_main:
        return "Только ДЗ"
    return "Есть пропуски"


def main() -> int:
    if not COURSE_ROOT.exists():
        print(f"Course root not found: {COURSE_ROOT}", file=sys.stderr)
        return 2
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    files, dirs, inaccessible = collect_tree()
    empty_dirs = [d for d in dirs if not any(d.iterdir())]

    by_topic = defaultdict(list)
    for item in files:
        by_topic[topic_key_for_file(item["path"])].append(item)

    topic_paths = sorted(set(by_topic) | set(empty_dirs), key=lambda p: str(p).lower())

    hash_groups = defaultdict(list)
    name_groups = defaultdict(list)
    for item in files:
        path = item["path"]
        if item["stat"].st_size > 0 and item["stat"].st_size <= 50 * 1024 * 1024:
            try:
                digest = file_hash(path)
                hash_groups[(digest, item["stat"].st_size)].append(path)
            except OSError:
                pass
        name_groups[(normalize_for_duplicate(path.name), item["stat"].st_size)].append(path)

    exact_dupes = {p for group in hash_groups.values() if len(group) > 1 for p in group}
    name_dupes = {p for group in name_groups.values() if len(group) > 1 for p in group}
    duplicate_paths = exact_dupes | name_dupes

    topic_rows = []
    file_rows = []
    problem_rows = []

    sibling_type_presence = defaultdict(Counter)
    temp_topic_flags = {}
    for topic_path in topic_paths:
        topic_files = by_topic.get(topic_path, [])
        flags = Counter()
        for item in topic_files:
            mt = material_type(item["path"])
            if mt in {"теория"}:
                flags["theory"] += 1
            if mt == "конспект":
                flags["conspect"] += 1
            if mt == "презентация":
                flags["presentation"] += 1
            if mt in {"вебинар", "запись урока"}:
                flags["video"] += 1
            if mt == "домашнее задание":
                flags["homework"] += 1
            if mt == "решение домашнего задания":
                flags["solution"] += 1
            if mt == "ответы":
                flags["answers"] += 1
            if mt in {"дополнительные задачи", "изображения и вспомогательные материалы", "архив", "другие учебные файлы"}:
                flags["extra"] += 1
        parent = topic_path.parent
        for key, value in flags.items():
            if value:
                sibling_type_presence[parent][key] += 1
        temp_topic_flags[topic_path] = flags

    topic_index = {path: idx + 1 for idx, path in enumerate(topic_paths)}

    for topic_path in topic_paths:
        meta = topic_meta(topic_path)
        topic_files = by_topic.get(topic_path, [])
        flags = temp_topic_flags[topic_path]
        sibling_count = len([p for p in topic_paths if p.parent == topic_path.parent])

        required = ["main", "homework", "solution"]
        if sibling_count >= 4 and sibling_type_presence[topic_path.parent]["answers"] >= max(2, sibling_count * 0.55):
            required.append("answers")

        missing = []
        if "main" in required and not (flags["theory"] or flags["conspect"] or flags["presentation"] or flags["video"]):
            missing.append("основной материал")
        if "homework" in required and not flags["homework"]:
            missing.append("ДЗ")
        if "solution" in required and not flags["solution"]:
            missing.append("решение ДЗ")
        if "answers" in required and not flags["answers"]:
            missing.append("ответы")

        uncertain = False
        if len(rel_parts(topic_path)) <= 1 and topic_files:
            uncertain = True
        if not extract_topic_number(topic_path.name) and len(topic_files) > 0 and len(rel_parts(topic_path)) >= 2:
            uncertain = True

        status = classify_topic_status(flags, len(topic_files), missing, uncertain)
        comment = []
        if "answers" not in required:
            comment.append("Ответы не считались обязательными, если это не следует из соседних тем.")
        if uncertain:
            comment.append("Номер или границы темы определены эвристически.")

        topic_rows.append(
            [
                topic_index[topic_path],
                meta["section"],
                meta["module"],
                meta["number"],
                meta["title"],
                str(topic_path),
                "Открыть папку",
                bool_yes(bool(flags["theory"] or flags["conspect"])),
                bool_yes(bool(flags["presentation"])),
                bool_yes(bool(flags["video"])),
                bool_yes(bool(flags["homework"])),
                bool_yes(bool(flags["solution"])),
                bool_yes(bool(flags["answers"])),
                bool_yes(bool(flags["extra"])),
                len(topic_files),
                ", ".join(missing),
                status,
                " ".join(comment),
            ]
        )

        if not topic_files:
            problem_rows.append(
                [
                    "Пустая папка",
                    topic_path.name,
                    str(topic_path),
                    "Открыть папку",
                    meta["title"],
                    "Папка не содержит файлов или вложенных объектов.",
                    "Проверить, нужна ли папка в курсе.",
                ]
            )
        if uncertain and topic_files:
            problem_rows.append(
                [
                    "Неоднозначная тема",
                    topic_path.name,
                    str(topic_path),
                    "Открыть папку",
                    meta["title"],
                    "Тему удалось определить только по папке, без явного номера/паттерна.",
                    "Проверить вручную название и место в структуре.",
                ]
            )

    for idx, item in enumerate(sorted(files, key=lambda it: str(it["path"]).lower()), start=1):
        path = item["path"]
        stat = item["stat"]
        topic_path = topic_key_for_file(path)
        meta = topic_meta(topic_path)
        mt = material_type(path)
        problem = []
        bind_status = "Привязан к теме"

        if path in duplicate_paths:
            bind_status = "Возможный дубль"
            problem.append("возможный дубль")
        if stat.st_size == 0:
            bind_status = "Файл повреждён или недоступен"
            problem.append("нулевой размер")
        if len(rel_parts(path)) <= 2 or not extract_topic_number(topic_path.name):
            if bind_status == "Привязан к теме":
                bind_status = "Привязан предположительно"
            problem.append("тема определена предположительно")
        if path.suffix.lower() in ARCHIVE_EXTS:
            problem.append("архив не раскрывался")

        file_rows.append(
            [
                idx,
                meta["section"],
                meta["module"],
                meta["number"],
                meta["title"],
                mt,
                path.name,
                path.suffix.lower(),
                str(path),
                "Открыть файл",
                stat.st_size,
                datetime.fromtimestamp(stat.st_mtime),
                bind_status,
                ", ".join(dict.fromkeys(problem)),
                "",
            ]
        )

        if path in duplicate_paths:
            problem_rows.append(
                [
                    "Возможный дубль",
                    path.name,
                    str(path),
                    "Открыть файл",
                    meta["title"],
                    "Найден файл с совпадающим хешем или нормализованным названием и размером.",
                    "Сравнить содержимое и оставить нужную версию.",
                ]
            )
        if stat.st_size == 0:
            problem_rows.append(
                [
                    "Пустой файл",
                    path.name,
                    str(path),
                    "Открыть файл",
                    meta["title"],
                    "Размер файла равен 0 байт.",
                    "Проверить источник файла или заменить.",
                ]
            )
        if path.suffix.lower() in ARCHIVE_EXTS:
            problem_rows.append(
                [
                    "Архив не проверен",
                    path.name,
                    str(path),
                    "Открыть файл",
                    meta["title"],
                    "Содержимое архива не раскрывалось в рамках безопасного аудита.",
                    "Открыть архив и проверить вложенные материалы отдельно.",
                ]
            )

    number_groups = defaultdict(list)
    for row in topic_rows:
        key = (row[1], row[2], row[3])
        if row[3]:
            number_groups[key].append(row)
    for (section, module, number), rows in number_groups.items():
        if len(rows) > 1:
            for row in rows:
                problem_rows.append(
                    [
                        "Повтор номера темы",
                        row[4],
                        row[5],
                        "Открыть папку",
                        row[4],
                        f"В разделе/модуле повторяется номер темы: {number}.",
                        "Проверить нумерацию и названия тем.",
                    ]
                )

    sorted_topics_by_parent = defaultdict(list)
    for path in topic_paths:
        number = extract_topic_number(path.name)
        if number and number.isdigit():
            sorted_topics_by_parent[path.parent].append((int(number), path))
    for parent, values in sorted_topics_by_parent.items():
        nums = sorted(n for n, _ in values)
        if len(nums) >= 3:
            expected = set(range(nums[0], nums[-1] + 1))
            missing_nums = sorted(expected - set(nums))
            if missing_nums and len(missing_nums) <= 20:
                problem_rows.append(
                    [
                        "Нарушение нумерации",
                        parent.name,
                        str(parent),
                        "Открыть папку",
                        "",
                        "В последовательности папок отсутствуют номера: " + ", ".join(map(str, missing_nums)),
                        "Проверить, не потеряны ли темы или не сбита ли нумерация.",
                    ]
                )

    for path, message in inaccessible:
        problem_rows.append(
            [
                "Файл недоступен",
                path.name,
                str(path),
                "Открыть файл",
                "",
                message,
                "Проверить права доступа или целостность файла.",
            ]
        )

    material_counter = Counter(row[5] for row in file_rows)
    material_topic_presence = defaultdict(set)
    for row in file_rows:
        material_topic_presence[row[5]].add((row[1], row[2], row[3], row[4]))

    summary_metrics = [
        ("Количество разделов", len(set(row[1] for row in topic_rows if row[1]))),
        ("Количество модулей", len(set((row[1], row[2]) for row in topic_rows if row[2]))),
        ("Количество тем", len(topic_rows)),
        ("Количество файлов", len(files)),
        ("Количество полностью заполненных тем", sum(1 for row in topic_rows if row[16] == "Полный комплект")),
        ("Количество тем с пропусками", sum(1 for row in topic_rows if row[16] == "Есть пропуски")),
        ("Количество пустых тем", sum(1 for row in topic_rows if row[16] == "Пустая тема")),
        ("Количество файлов без привязки", sum(1 for row in file_rows if row[12] == "Не удалось определить тему")),
        ("Количество возможных дублей", len(duplicate_paths)),
        ("Количество пустых папок", len(empty_dirs)),
        ("Количество нарушений нумерации", sum(1 for row in problem_rows if row[0] == "Нарушение нумерации")),
    ]

    wb = Workbook()
    ws_topics = wb.active
    ws_topics.title = "Темы курса"
    ws_files = wb.create_sheet("Все файлы")
    ws_missing = wb.create_sheet("Пропуски")
    ws_problems = wb.create_sheet("Проблемные файлы")
    ws_summary = wb.create_sheet("Сводка")

    headers_topics = [
        "№",
        "Раздел",
        "Модуль",
        "Номер темы",
        "Название темы",
        "Папка темы",
        "Ссылка на папку",
        "Теория",
        "Презентация",
        "Вебинар / видео",
        "Домашнее задание",
        "Решение ДЗ",
        "Ответы",
        "Дополнительные материалы",
        "Количество файлов",
        "Что отсутствует",
        "Статус темы",
        "Комментарий",
    ]
    headers_files = [
        "№",
        "Раздел",
        "Модуль",
        "Номер темы",
        "Название темы",
        "Тип материала",
        "Название файла",
        "Расширение",
        "Полный путь",
        "Гиперссылка на файл",
        "Размер файла",
        "Дата изменения",
        "Статус привязки",
        "Возможная проблема",
        "Комментарий",
    ]
    headers_missing = [
        "Раздел",
        "Модуль",
        "Номер темы",
        "Название темы",
        "Отсутствует теория",
        "Отсутствует презентация",
        "Отсутствует видео",
        "Отсутствует ДЗ",
        "Отсутствует решение ДЗ",
        "Отсутствуют ответы",
        "Папка темы",
        "Ссылка на папку",
        "Комментарий",
    ]
    headers_problems = [
        "Тип проблемы",
        "Название объекта",
        "Полный путь",
        "Гиперссылка",
        "Предполагаемая тема",
        "Описание проблемы",
        "Рекомендуемое действие",
    ]

    def write_sheet(ws, headers, rows):
        ws.append(headers)
        for row in rows:
            ws.append(row)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="305496")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    missing_rows = []
    for row in topic_rows:
        if row[15] or row[16] in {"Есть пропуски", "Только теория", "Только ДЗ", "Пустая тема", "Требует ручной проверки"}:
            missing_text = row[15]
            missing_rows.append(
                [
                    row[1],
                    row[2],
                    row[3],
                    row[4],
                    bool_yes("основной материал" in missing_text),
                    bool_yes(row[8] == "Нет"),
                    bool_yes(row[9] == "Нет"),
                    bool_yes("ДЗ" in missing_text),
                    bool_yes("решение ДЗ" in missing_text),
                    bool_yes("ответы" in missing_text),
                    row[5],
                    "Открыть папку",
                    row[17],
                ]
            )

    write_sheet(ws_topics, headers_topics, topic_rows)
    write_sheet(ws_files, headers_files, file_rows)
    write_sheet(ws_missing, headers_missing, missing_rows)
    write_sheet(ws_problems, headers_problems, problem_rows)

    ws_summary.append(["Показатель", "Значение"])
    for metric in summary_metrics:
        ws_summary.append(list(metric))
    ws_summary.append([])
    ws_summary.append(["Тип материала", "Найдено файлов", "В скольких темах есть", "В скольких темах отсутствует"])
    total_topics = len(topic_rows)
    for mt, count in sorted(material_counter.items()):
        present = len(material_topic_presence[mt])
        ws_summary.append([mt, count, present, max(0, total_topics - present)])
    ws_summary.freeze_panes = "A2"
    ws_summary.auto_filter.ref = ws_summary.dimensions
    for cell in ws_summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="305496")
    for cell in ws_summary[14]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="70AD47")

    # Hyperlinks.
    for r in range(2, ws_topics.max_row + 1):
        folder = Path(ws_topics.cell(r, 6).value)
        ws_topics.cell(r, 7).hyperlink = win_link(folder)
        ws_topics.cell(r, 7).style = "Hyperlink"
    for r in range(2, ws_files.max_row + 1):
        file_path = Path(ws_files.cell(r, 9).value)
        ws_files.cell(r, 10).hyperlink = win_link(file_path)
        ws_files.cell(r, 10).style = "Hyperlink"
        ws_files.cell(r, 12).number_format = "yyyy-mm-dd hh:mm"
    for r in range(2, ws_missing.max_row + 1):
        folder = Path(ws_missing.cell(r, 11).value)
        ws_missing.cell(r, 12).hyperlink = win_link(folder)
        ws_missing.cell(r, 12).style = "Hyperlink"
    for r in range(2, ws_problems.max_row + 1):
        target = Path(ws_problems.cell(r, 3).value)
        ws_problems.cell(r, 4).hyperlink = win_link(target)
        ws_problems.cell(r, 4).style = "Hyperlink"

    green = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red = PatternFill("solid", fgColor="FFC7CE")
    for ws in [ws_topics, ws_files, ws_missing, ws_problems, ws_summary]:
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = 0
            for cell in ws[letter]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 70))
            ws.column_dimensions[letter].width = max(10, min(max_len + 2, 55))
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 18 if row == 1 else 34

    status_col = 17
    for r in range(2, ws_topics.max_row + 1):
        cell = ws_topics.cell(r, status_col)
        if cell.value == "Полный комплект":
            cell.fill = green
        elif cell.value in {"Требует ручной проверки", "Привязан предположительно"}:
            cell.fill = yellow
        elif cell.value:
            cell.fill = red

    for r in range(2, ws_files.max_row + 1):
        cell = ws_files.cell(r, 13)
        if cell.value == "Привязан к теме":
            cell.fill = green
        elif cell.value == "Привязан предположительно":
            cell.fill = yellow
        else:
            cell.fill = red

    for ws in [ws_missing, ws_problems]:
        ws.conditional_formatting.add(f"A2:{get_column_letter(ws.max_column)}{ws.max_row}", CellIsRule(operator="notEqual", formula=['""'], fill=None))

    status_validation = DataValidation(
        type="list",
        formula1='"Полный комплект,Есть пропуски,Только теория,Только ДЗ,Пустая тема,Требует ручной проверки"',
        allow_blank=True,
    )
    ws_topics.add_data_validation(status_validation)
    status_validation.add(f"Q2:Q{max(2, ws_topics.max_row)}")

    wb.save(XLSX_PATH)

    loaded = load_workbook(XLSX_PATH, read_only=False, data_only=False)
    expected_sheets = {"Темы курса", "Все файлы", "Пропуски", "Проблемные файлы", "Сводка"}
    assert expected_sheets <= set(loaded.sheetnames)
    assert loaded["Все файлы"].max_row - 1 == len(files)
    assert loaded["Темы курса"]["G2"].hyperlink is not None if loaded["Темы курса"].max_row >= 2 else True
    loaded.close()

    critical = [
        row for row in topic_rows
        if row[16] in {"Пустая тема", "Только теория", "Только ДЗ", "Требует ручной проверки"}
    ][:25]

    md_lines = [
        "# Аудит материалов курса",
        "",
        f"Проверенная папка: `{COURSE_ROOT}`",
        f"Дата аудита: {datetime.now():%Y-%m-%d %H:%M}",
        "",
        "## Сводка",
        "",
    ]
    for name, value in summary_metrics:
        md_lines.append(f"- {name}: {value}")
    md_lines.extend(
        [
            "",
            "## Правила распознавания",
            "",
            "- Темой считалась папка, в которой непосредственно лежат учебные файлы, а также пустая папка, если она обнаружена в дереве курса.",
            "- Раздел определялся по первой папке внутри корня курса; модуль - по второй папке, если она есть.",
            "- Номер темы извлекался из начала имени папки, конструкций `№...`, `Вариант №...`, `Вебинар ...`, `ДЗ-...`.",
            "- Тип материала определялся по названию файла, названиям ближайших папок и расширению.",
            "- Возможные дубли определялись по совпадению SHA1+размера и по нормализованному имени+размеру.",
            "",
            "## Обязательные материалы",
            "",
            "- Основной материал: теория, конспект, презентация, вебинар или видео.",
            "- Домашнее задание.",
            "- Решение или разбор домашнего задания.",
            "- Ответы считались обязательными только там, где такая структура заметна у большинства соседних тем.",
            "",
            "## Неоднозначности",
            "",
            "- В структуре есть папки, где тема совпадает с вебинаром, вариантом или номером задания ЕГЭ, поэтому часть тем помечена как требующая ручной проверки.",
            "- Архивы не раскрывались, чтобы не изменять исходные материалы и не создавать временные файлы внутри курса.",
            "- Материалы верхнего уровня без явной папки-темы привязаны предположительно.",
            "",
            "## Самые критичные пропуски / ручная проверка",
            "",
        ]
    )
    if critical:
        for row in critical:
            md_lines.append(f"- [{row[16]}] {row[1]} / {row[2]} / {row[4]} - отсутствует: {row[15] or 'требует проверки'}")
    else:
        md_lines.append("- Критичных пропусков по выбранным эвристикам не найдено.")
    md_lines.extend(
        [
            "",
            "## Итоговые файлы",
            "",
            f"1. Excel-файл: `{XLSX_PATH}`",
            f"2. Технический отчет: `{MD_PATH}`",
        ]
    )
    MD_PATH.write_text("\n".join(md_lines), encoding="utf-8")

    print(f"XLSX={XLSX_PATH}")
    print(f"MD={MD_PATH}")
    print(f"FILES={len(files)}")
    print(f"DIRS={len(dirs)}")
    print(f"TOPICS={len(topic_rows)}")
    print(f"PROBLEMS={len(problem_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
